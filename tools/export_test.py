# -*- coding: utf-8 -*-
"""Testa o export Excel (todas as peças)."""
import os, sys, io
sys.stdout.reconfigure(encoding="utf-8")
APP = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "app")
sys.path.insert(0, APP)
import configs as C, ingest as I, extrato as E, export as EXP

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
cfg = C.load(os.path.join(ROOT, "config"))
dfs = I.load_all(ROOT, cfg)
tx, rs = E.load_transacoes(ROOT)
per = sorted(dfs["periodos"])
cx, _ = E.caixa_real_fim(rs, per[-1])
adi, apo = E.passivo_pl_extras(tx, per[-1])

data = EXP.build_excel(dfs, per, cfg, True, cx, adi, apo, tx)
out = os.path.join(ROOT, "Painel_3Amores_Acumulado.xlsx")
with open(out, "wb") as f: f.write(data)
print(f"✓ Excel exportado: {os.path.basename(out)}  ({len(data)//1024} KB)")
import openpyxl
wb = openpyxl.load_workbook(io.BytesIO(data))
print("  Abas:", ", ".join(wb.sheetnames))
for sh in wb.sheetnames:
    print(f"    - {sh}: {wb[sh].max_row} linhas")
