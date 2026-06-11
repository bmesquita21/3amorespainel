# -*- coding: utf-8 -*-
"""Gera UMA planilha única e limpa p/ a Sabrina classificar: descrições enriquecidas (pagador),
preserva rótulos já preenchidos nos arquivos antigos, coluna 'sugestão', e move os antigos p/ subpasta."""
import os, sys, glob, shutil
sys.stdout.reconfigure(encoding="utf-8")
APP = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "app")
sys.path.insert(0, APP)
import pandas as pd
import configs as C, ingest as I, extrato as E, brutils as B

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DESK = r"C:\Users\Sabrina\OneDrive - Grupo Bom Jardim\Área de Trabalho"
cfg = C.load(os.path.join(ROOT, "config"))
dfs = I.load_all(ROOT, cfg)
tx, rs = E.load_transacoes(ROOT)
outros = E.creditos_outros(tx)   # tid, periodo, data, banco, titular, conta, valor, desc, natureza (preencher)

# 1) RECUPERA rótulos já preenchidos em qualquer planilha antiga (Desktop + projeto)
def _labels_de(fp):
    out = {}
    try: df = pd.read_excel(fp)
    except Exception: return out
    cols = {str(c).strip().lower(): c for c in df.columns}
    ncol = next((cols[k] for k in cols if k.startswith("natureza")), None)
    if not ncol: return out
    kcol = cols.get("tid") or cols.get("id")
    bcol, ccol, dcol, vcol = cols.get("banco"), cols.get("conta"), cols.get("data"), cols.get("valor")
    for _, r in df.iterrows():
        nat = str(r[ncol]).strip()
        if not nat or nat.lower() == "nan": continue
        if kcol and str(r[kcol]).strip() and str(r[kcol]).strip().lower() != "nan":
            out[str(r[kcol]).strip()] = nat
        elif bcol and ccol and dcol and vcol:
            dv = r[dcol]; ds = dv.strftime("%d/%m/%Y") if hasattr(dv, "strftime") else str(dv).strip()
            try: vv = float(r[vcol])
            except Exception: vv = float(str(r[vcol]).replace("R$", "").replace(".", "").replace(",", ".").strip() or 0)
            out[E._tid(str(r[bcol]).strip(), str(r[ccol]).strip(), ds, vv)] = nat
    return out

antigos = glob.glob(os.path.join(DESK, "CREDITOS*.xlsx")) + glob.glob(os.path.join(ROOT, "creditos_outros*.xlsx"))
labels = {}
for fp in antigos:
    labels.update(_labels_de(fp))
print(f"rótulos já preenchidos recuperados: {len(labels)}")

# 2) MONTA a planilha única
df = outros.copy()
df["natureza (preencher)"] = df["tid"].map(lambda t: labels.get(t, ""))
def sug(d):
    n = E._norm(d)
    return "aporte" if ("ALVARO" in n or "FREITAS" in n) else ""
df["sugestão (confira)"] = df["desc"].map(sug)
df = df.rename(columns={"tid": "id", "periodo": "Mês", "data": "Data", "banco": "Banco",
                        "titular": "Titular", "conta": "Conta", "valor": "Valor (R$)", "desc": "Histórico / Pagador"})
df = df[["Mês", "Data", "Banco", "Titular", "Conta", "Valor (R$)", "Histórico / Pagador",
         "sugestão (confira)", "natureza (preencher)", "id"]]
nsug = int((df["sugestão (confira)"] != "").sum()); npre = int((df["natureza (preencher)"] != "").sum())
tot = df["Valor (R$)"].sum(); top50 = df.head(50)["Valor (R$)"].sum()
print(f"linhas: {len(df)} | já preenchidas: {npre} | sugestões (Álvaro): {nsug}")
print(f"total Outros: {B.brl(tot)} | top 50 = {B.brl(top50)} ({100*top50/tot:.0f}% do valor)")

# 3) ESCREVE com formatação
OUT = os.path.join(DESK, "CREDITOS_classificar.xlsx")
with pd.ExcelWriter(OUT, engine="openpyxl") as xl:
    df.to_excel(xl, index=False, sheet_name="Classificar")
    ws = xl.sheets["Classificar"]
    from openpyxl.styles import Font, PatternFill, Alignment
    hdr = PatternFill("solid", fgColor="D9E1F2"); ylw = PatternFill("solid", fgColor="FFF2CC")
    for c in ws[1]:
        c.font = Font(bold=True); c.fill = hdr; c.alignment = Alignment(horizontal="center")
    widths = {"A": 9, "B": 12, "C": 11, "D": 9, "E": 13, "F": 15, "G": 52, "H": 18, "I": 22, "J": 13}
    for col, w in widths.items(): ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    for row in range(2, len(df) + 2):
        ws[f"F{row}"].number_format = '#,##0.00'
        ws[f"I{row}"].fill = ylw   # coluna 'natureza (preencher)' destacada
print(f"✓ planilha única: {OUT}")

# 4) MOVE os antigos do Desktop p/ subpasta (não apaga)
bak = os.path.join(DESK, "_planilhas_antigas")
os.makedirs(bak, exist_ok=True)
for f in ("CREDITOS_OUTROS_classificar.xlsx", "CREDITOS_OUTROS_v2_classificar.xlsx"):
    p = os.path.join(DESK, f)
    if os.path.exists(p):
        shutil.move(p, os.path.join(bak, f)); print(f"  movido p/ _planilhas_antigas: {f}")
